"""
Procesador de Transacciones Bancarias - SIN PANDAS
Archivo: src/business/processor.py
"""

import csv
import re
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from src.business.categorizer import TransactionCategorizer


class TransactionProcessor:
    """Clase para procesar y limpiar archivos de transacciones bancarias"""

    def __init__(self):
        self.categorizer = TransactionCategorizer()
        self.data = []  # Lista de diccionarios en lugar de DataFrame
        self.errors = []
        self.original_count = 0

    def load_file(self, file_path: str) -> Tuple[bool, str]:
        """
        Carga un archivo CSV o Excel
        Returns: (success, message)
        """
        try:
            self.errors = []
            self.data = []

            if file_path.endswith(".csv"):
                # Intentar diferentes encodings
                encodings = ["utf-8", "latin-1", "iso-8859-1", "cp1252"]
                for encoding in encodings:
                    try:
                        with open(file_path, 'r', encoding=encoding) as f:
                            reader = csv.DictReader(f)
                            self.data = list(reader)
                        break
                    except UnicodeDecodeError:
                        continue
                else:
                    return False, "No se pudo decodificar el archivo CSV"

            elif file_path.endswith((".xlsx", ".xls")):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(file_path, read_only=True)
                    ws = wb.active
                    
                    # Leer encabezados
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    
                    # Leer datos
                    self.data = []
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        row_dict = {headers[i]: row[i] for i in range(len(headers)) if i < len(row)}
                        self.data.append(row_dict)
                    
                    wb.close()
                except ImportError:
                    return False, "openpyxl no está disponible para leer Excel"
            else:
                return False, "Formato no soportado. Use CSV o Excel (.xlsx, .xls)"

            if not self.data:
                return False, "El archivo está vacío"

            self.original_count = len(self.data)
            return True, f"Archivo cargado: {self.original_count} registros"

        except FileNotFoundError:
            return False, "Archivo no encontrado"
        except Exception as e:
            return False, f"Error al cargar archivo: {str(e)}"

    def validate_columns(self) -> Tuple[bool, str]:
        """
        Valida que el archivo tenga las columnas necesarias
        """
        if not self.data:
            return False, "No hay datos cargados"

        # Obtener columnas del primer registro
        columns = list(self.data[0].keys())
        columns_lower = [col.lower().strip() for col in columns]

        # Buscar columnas
        date_keywords = ["fecha", "date", "dia", "day", "when", "datetime"]
        desc_keywords = ["descripcion", "description", "concepto", "detalle", "detail", "desc", "memo", "nota", "note"]
        amount_keywords = ["monto", "amount", "importe", "valor", "value", "precio", "price", "total", "cantidad", "quantity"]
        type_keywords = ["tipo", "type", "categoria", "category"]

        date_cols = [col for col in columns_lower if any(kw in col for kw in date_keywords)]
        desc_cols = [col for col in columns_lower if any(kw in col for kw in desc_keywords)]
        amount_cols = [col for col in columns_lower if any(kw in col for kw in amount_keywords)]
        type_cols = [col for col in columns_lower if any(kw in col for kw in type_keywords)]

        # Validaciones
        if not date_cols:
            return False, f"No se encontró columna de fecha. Columnas: {', '.join(columns)}"
        if not desc_cols:
            return False, f"No se encontró columna de descripción. Columnas: {', '.join(columns)}"
        if not amount_cols:
            return False, f"No se encontró columna de monto. Columnas: {', '.join(columns)}"

        # Renombrar columnas
        original_date = columns[columns_lower.index(date_cols[0])]
        original_desc = columns[columns_lower.index(desc_cols[0])]
        original_amount = columns[columns_lower.index(amount_cols[0])]
        
        rename_map = {
            original_date: "fecha",
            original_desc: "descripcion",
            original_amount: "monto"
        }
        
        if type_cols:
            original_type = columns[columns_lower.index(type_cols[0])]
            rename_map[original_type] = "tipo"

        # Aplicar renombrado
        new_data = []
        for row in self.data:
            new_row = {}
            for old_key, value in row.items():
                new_key = rename_map.get(old_key, old_key)
                new_row[new_key] = value
            new_data.append(new_row)
        
        self.data = new_data

        message = f"Columnas validadas: fecha='{original_date}', descripcion='{original_desc}', monto='{original_amount}'"
        if type_cols:
            message += f", tipo='{original_type}'"

        return True, message

    def _parse_date(self, date_str) -> Optional[datetime]:
        """Intenta parsear una fecha en múltiples formatos"""
        if not date_str or str(date_str).strip() == "":
            return None
        
        date_str = str(date_str).strip()
        
        formats = [
            "%Y-%m-%d",
            "%d/%m/%Y",
            "%m/%d/%Y",
            "%Y/%m/%d",
            "%d-%m-%Y",
            "%Y-%m-%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
        ]
        
        for fmt in formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        return None

    def _parse_amount(self, amount_str) -> Optional[float]:
        """Limpia y convierte un monto a float"""
        if not amount_str:
            return None
        
        amount_str = str(amount_str)
        
        # Eliminar símbolos de moneda
        currency_symbols = ["S/", "S/.", "$", "€", "£", "USD", "PEN", "EUR"]
        for symbol in currency_symbols:
            amount_str = amount_str.replace(symbol, "")
        
        # Eliminar espacios y comas
        amount_str = amount_str.replace(" ", "").replace(",", "")
        
        try:
            return float(amount_str)
        except ValueError:
            return None

    def clean_data(self) -> Tuple[bool, str]:
        """
        Limpia y estandariza los datos
        """
        if not self.data:
            return False, "No hay datos para limpiar"

        initial_count = len(self.data)
        self.errors = []
        cleaned_data = []

        for row in self.data:
            # Eliminar filas completamente vacías
            if not any(row.values()):
                continue
            
            # Parsear fecha
            date_obj = self._parse_date(row.get("fecha"))
            if not date_obj:
                self.errors.append(f"⚠️ Fecha inválida eliminada: {row.get('fecha')}")
                continue
            
            # Limpiar descripción
            desc = str(row.get("descripcion", "")).strip()
            if desc.lower() in ["", "nan", "none", "null", "n/a", "na"]:
                self.errors.append(f"⚠️ Descripción vacía eliminada")
                continue
            
            # Parsear monto
            amount = self._parse_amount(row.get("monto"))
            if amount is None:
                self.errors.append(f"⚠️ Monto inválido eliminado: {row.get('monto')}")
                continue
            
            # Convertir a positivo
            amount = abs(amount)
            
            # Eliminar montos muy pequeños
            if amount <= 0.001:
                self.errors.append(f"⚠️ Monto en cero eliminado")
                continue
            
            # Procesar columna tipo
            tipo = str(row.get("tipo", "expense")).lower().strip()
            type_mapping = {
                "gasto": "expense", "gastos": "expense", "expense": "expense",
                "egreso": "expense", "egresos": "expense", "salida": "expense",
                "ingreso": "income", "ingresos": "income", "income": "income",
                "entrada": "income"
            }
            tipo = type_mapping.get(tipo, "expense")
            
            # Crear registro limpio
            cleaned_row = {
                "fecha": date_obj,
                "descripcion": desc,
                "monto": amount,
                "tipo": tipo
            }
            
            cleaned_data.append(cleaned_row)

        # Eliminar duplicados
        unique_data = []
        seen = set()
        for row in cleaned_data:
            key = (row["fecha"], row["descripcion"], row["monto"])
            if key not in seen:
                seen.add(key)
                unique_data.append(row)
        
        duplicates = len(cleaned_data) - len(unique_data)
        if duplicates > 0:
            self.errors.append(f"⚠️ {duplicates} registros duplicados eliminados")
        
        # Ordenar por fecha (más reciente primero)
        unique_data.sort(key=lambda x: x["fecha"], reverse=True)
        
        self.data = unique_data
        final_count = len(self.data)
        removed = initial_count - final_count

        message = f"✅ Limpieza completada: {final_count} registros válidos"
        if removed > 0:
            message += f" ({removed} registros eliminados)"
        if self.errors and len(self.errors) <= 10:
            message += "\n" + "\n".join(self.errors[:10])

        return True, message

    def categorize_transactions(self, categories_map_expense: Dict[int, str], 
                                categories_map_income: Dict[int, str]) -> bool:
        """
        Asigna categorías a las transacciones
        """
        if not self.data:
            return False

        try:
            # Invertir mapas
            name_to_id_expense = {str(v).strip().lower(): int(k) for k, v in categories_map_expense.items()}
            name_to_id_income = {str(v).strip().lower(): int(k) for k, v in categories_map_income.items()}

            for row in self.data:
                desc = row.get("descripcion", "").lower()
                transaction_type = row.get("tipo", "expense")
                
                # Categorizar
                category_name = self.categorizer.categorize(desc, transaction_type)
                category_name_lower = category_name.lower().strip()
                
                # Elegir mapa correcto
                if transaction_type == "income":
                    name_to_id = name_to_id_income
                    default_id = next(iter(name_to_id_income.values())) if name_to_id_income else 1
                else:
                    name_to_id = name_to_id_expense
                    default_id = name_to_id.get("otros gastos", 
                                 name_to_id.get("otros", 
                                 next(iter(name_to_id.values())) if name_to_id else 1))
                
                category_id = name_to_id.get(category_name_lower, default_id)
                row["categoria_id"] = category_id

            return True

        except Exception as e:
            print(f"❌ Error en categorización: {e}")
            return False

    def get_processed_data(self) -> List[Dict]:
        """
        Retorna los datos procesados listos para insertar en la BD
        """
        if not self.data:
            return []

        transactions = []
        for row in self.data:
            transactions.append({
                "date": row["fecha"],
                "description": row["descripcion"],
                "amount": float(row["monto"]),
                "category_id": int(row["categoria_id"]),
                "transaction_type": row["tipo"],
                "source": "imported",
                "original_description": row["descripcion"],
            })

        return transactions

    def get_summary(self) -> Dict:
        """Obtiene un resumen de los datos procesados"""
        if not self.data:
            return {"status": "empty", "message": "No hay datos procesados"}

        try:
            expenses = [r for r in self.data if r["tipo"] == "expense"]
            incomes = [r for r in self.data if r["tipo"] == "income"]
            
            total_expenses = sum(r["monto"] for r in expenses)
            total_income = sum(r["monto"] for r in incomes)
            amounts = [r["monto"] for r in self.data]
            dates = [r["fecha"] for r in self.data]

            summary = {
                "status": "success",
                "original_count": self.original_count,
                "processed_count": len(self.data),
                "total_transactions": len(self.data),
                "removed_count": self.original_count - len(self.data),
                "total_amount": sum(amounts),
                "total_expenses": total_expenses,
                "total_income": total_income,
                "count_expenses": len(expenses),
                "count_income": len(incomes),
                "average_amount": sum(amounts) / len(amounts) if amounts else 0,
                "min_amount": min(amounts) if amounts else 0,
                "max_amount": max(amounts) if amounts else 0,
                "date_range": {
                    "start": min(dates).strftime("%Y-%m-%d") if dates else "",
                    "end": max(dates).strftime("%Y-%m-%d") if dates else "",
                },
                "errors": self.errors,
            }

            return summary

        except Exception as e:
            return {"status": "error", "message": f"Error: {str(e)}"}

    def preview_data(self, rows: int = 10) -> Optional[List[Dict]]:
        """Retorna una vista previa de los datos procesados"""
        if not self.data:
            return None
        return self.data[:rows]

    def export_to_dict(self) -> List[Dict]:
        """Exporta los datos procesados a una lista de diccionarios"""
        return self.data if self.data else []

    def reset(self):
        """Resetea el procesador"""
        self.data = []
        self.errors = []
        self.original_count = 0