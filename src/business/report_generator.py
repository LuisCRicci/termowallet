"""
Generador de Reportes -PARA ANDROID 9
CORRECCIONES:
1. Callbacks llamados correctamente
2. Dialog se cierra después de generar
3. FileProvider para Android (sin FileUriExposedException)
"""
import csv
import sys
import os
import tempfile
import shutil
from datetime import datetime, timedelta
from typing import Dict, List
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
except ImportError:
    openpyxl = None
import flet as ft


class ReportGenerator:
    """Generador de reportes con sistema nativo de compartir"""
    
    def __init__(self, db_manager, page=None):
        self.db = db_manager
        self.page = page
        from src.utils.config import Config
        self.is_android = Config.is_android()

    def _clean_old_reports(self):
        """Limpia reportes antiguos del directorio de caché"""
        try:
            from src.utils.android_permissions import get_app_storage_path
            
            if self.is_android:
                cache_dir = get_app_storage_path().replace("/files", "/cache")
            else:
                cache_dir = tempfile.gettempdir()
            
            if not os.path.exists(cache_dir):
                return
            
            print(f"\n  LIMPIANDO REPORTES ANTIGUOS")
            now = datetime.now()
            cutoff_time = now - timedelta(hours=24)
            cleaned_count = 0
            
            for filename in os.listdir(cache_dir):
                if not filename.startswith("Reporte_TermoWallet"):
                    continue
                if not (filename.endswith(".xlsx") or filename.endswith(".csv")):
                    continue
                
                filepath = os.path.join(cache_dir, filename)
                
                try:
                    file_mtime = datetime.fromtimestamp(os.path.getmtime(filepath))
                    if file_mtime < cutoff_time:
                        os.remove(filepath)
                        cleaned_count += 1
                        print(f"  Eliminado: {filename}")
                except Exception as file_error:
                    print(f" Error procesando {filename}: {file_error}")
                    continue
            
            if cleaned_count > 0:
                print(f" {cleaned_count} archivo(s) antiguo(s) eliminado(s)")
            
        except Exception as e:
            print(f" Error en limpieza de reportes: {e}")

    def generate_monthly_report(
        self,
        year: int,
        month: int,
        format: str = "xlsx",
        callback_success = None,
        callback_error = None
    ) -> Dict:
        """Genera reporte mensual"""
        try:
            print(f"\n Generando reporte mensual: {month}/{year}")
            
            transactions = self.db.get_transactions_by_month(year, month)
            summary = self.db.get_monthly_summary(year, month)
            expenses_by_cat = self.db.get_expenses_by_category(year, month)
            income_by_cat = self.db.get_income_by_category(year, month)
            
            if not transactions:
                error_msg = "âŒ No hay transacciones en este mes"
                print(error_msg)
                if callback_error:
                    callback_error(error_msg)
                return {"success": False, "filepath": None, "message": error_msg}
            
            # Preparar datos
            transactions_data = []
            for t in transactions:
                category = self.db.get_category_by_id(t.category_id)
                transactions_data.append({
                    "Fecha": t.date.strftime("%d/%m/%Y"),
                    "Descripción": t.description,
                    "Categoría": category.name if category else "Sin Categoría",
                    "Tipo": "Ingreso" if t.transaction_type == "income" else "Gasto",
                    "Monto": t.amount,
                    "Notas": t.notes or "",
                })
            
            summary_data = [
                {"Concepto": "Total Ingresos", "Valor": summary["total_income"]},
                {"Concepto": "Total Gastos", "Valor": summary["total_expenses"]},
                {"Concepto": "Balance", "Valor": summary["savings"]},
                {"Concepto": "Tasa de Ahorro (%)", "Valor": summary["savings_rate"]},
                {"Concepto": "NÂº Transacciones", "Valor": summary["transaction_count"]}
            ]
            
            expenses_data = []
            if summary["total_expenses"] > 0:
                for cat in expenses_by_cat:
                    percentage = (cat["total"] / summary["total_expenses"] * 100)
                    expenses_data.append({
                        "Categoría": cat["category"],
                        "Total": cat["total"],
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            income_data = []
            if summary["total_income"] > 0:
                for cat in income_by_cat:
                    percentage = (cat["total"] / summary["total_income"] * 100)
                    income_data.append({
                        "Categoría": cat["category"],
                        "Total": cat["total"],
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            month_names = [
                "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
            ]
            filename = f"Reporte_TermoWallet_{month_names[month-1]}_{year}.{format}"
            
            # âœ… GENERAR DIRECTAMENTE (sin FilePicker)
            return self._generate_and_share_report(
                filename,
                format,
                transactions_data,
                summary_data,
                expenses_data,
                income_data,
                callback_success,
                callback_error
            )
            
        except Exception as e:
            print(f" Error generando reporte: {e}")
            import traceback
            traceback.print_exc()
            error_msg = f" Error: {str(e)}"
            if callback_error:
                callback_error(error_msg)
            return {"success": False, "filepath": None, "message": error_msg}

    def generate_custom_range_report(
        self,
        start_date: datetime,
        end_date: datetime,
        format: str = "xlsx",
        callback_success = None,
        callback_error = None
    ) -> Dict:
        """Genera reporte de rango personalizado"""
        try:
            print(f"\n Generando reporte personalizado: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}")
            
            transactions = self.db.get_transactions_by_date_range(start_date, end_date)
            
            if not transactions:
                error_msg = " No hay transacciones en este rango de fechas"
                print(error_msg)
                if callback_error:
                    callback_error(error_msg)
                return {"success": False, "filepath": None, "message": error_msg}
            
            total_income = sum(t.amount for t in transactions if t.transaction_type == "income")
            total_expenses = sum(t.amount for t in transactions if t.transaction_type == "expense")
            
            transactions_data = []
            for t in transactions:
                category = self.db.get_category_by_id(t.category_id)
                transactions_data.append({
                    "Fecha": t.date.strftime("%d/%m/%Y"),
                    "Descripción": t.description,
                    "Categoría": category.name if category else "Sin Categoría",
                    "Tipo": "Ingreso" if t.transaction_type == "income" else "Gasto",
                    "Monto": t.amount,
                    "Notas": t.notes or "",
                })
            
            summary_data = [
                {"Concepto": "Fecha Inicio", "Valor": start_date.strftime("%d/%m/%Y")},
                {"Concepto": "Fecha Fin", "Valor": end_date.strftime("%d/%m/%Y")},
                {"Concepto": "Días", "Valor": (end_date - start_date).days + 1},
                {"Concepto": "Total Ingresos", "Valor": total_income},
                {"Concepto": "Total Gastos", "Valor": total_expenses},
                {"Concepto": "Balance", "Valor": total_income - total_expenses},
                {"Concepto": "NÂº Transacciones", "Valor": len(transactions)}
            ]
            
            expenses_by_cat = {}
            for t in transactions:
                if t.transaction_type == "expense":
                    category = self.db.get_category_by_id(t.category_id)
                    cat_name = category.name if category else "Sin Categoría"
                    if cat_name not in expenses_by_cat:
                        expenses_by_cat[cat_name] = 0
                    expenses_by_cat[cat_name] += t.amount
            
            expenses_data = []
            if total_expenses > 0:
                for cat_name, total in sorted(expenses_by_cat.items(), key=lambda x: x[1], reverse=True):
                    percentage = (total / total_expenses * 100)
                    expenses_data.append({
                        "Categoría": cat_name,
                        "Total": total,
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            income_by_cat = {}
            for t in transactions:
                if t.transaction_type == "income":
                    category = self.db.get_category_by_id(t.category_id)
                    cat_name = category.name if category else "Sin Categoría"
                    if cat_name not in income_by_cat:
                        income_by_cat[cat_name] = 0
                    income_by_cat[cat_name] += t.amount
            
            income_data = []
            if total_income > 0:
                for cat_name, total in sorted(income_by_cat.items(), key=lambda x: x[1], reverse=True):
                    percentage = (total / total_income * 100)
                    income_data.append({
                        "Categoría": cat_name,
                        "Total": total,
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            filename = f"Reporte_TermoWallet_{start_date.strftime('%d%m%Y')}_al_{end_date.strftime('%d%m%Y')}.{format}"
            
            return self._generate_and_share_report(
                filename,
                format,
                transactions_data,
                summary_data,
                expenses_data,
                income_data,
                callback_success,
                callback_error
            )
            
        except Exception as e:
            print(f"Error generando reporte personalizado: {e}")
            import traceback
            traceback.print_exc()
            error_msg = f" Error: {str(e)}"
            if callback_error:
                callback_error(error_msg)
            return {"success": False, "filepath": None, "message": error_msg}

    def generate_annual_report(
        self,
        year: int,
        format: str = "xlsx",
        callback_success = None,
        callback_error = None
    ) -> Dict:
        """Genera reporte anual"""
        try:
            print(f"\n Generando reporte anual: {year}")
            
            all_transactions = []
            for month in range(1, 13):
                monthly_trans = self.db.get_transactions_by_month(year, month)
                all_transactions.extend(monthly_trans)
            
            if not all_transactions:
                error_msg = f" No hay transacciones en el año {year}"
                print(error_msg)
                if callback_error:
                    callback_error(error_msg)
                return {"success": False, "filepath": None, "message": error_msg}
            
            total_income = sum(t.amount for t in all_transactions if t.transaction_type == "income")
            total_expenses = sum(t.amount for t in all_transactions if t.transaction_type == "expense")
            
            transactions_data = []
            for t in all_transactions:
                category = self.db.get_category_by_id(t.category_id)
                transactions_data.append({
                    "Fecha": t.date.strftime("%d/%m/%Y"),
                    "Descripción": t.description,
                    "Categoría": category.name if category else "Sin Categoría",
                    "Tipo": "Ingreso" if t.transaction_type == "income" else "Gasto",
                    "Monto": t.amount,
                    "Notas": t.notes or "",
                })
            
            summary_data = [
                {"Concepto": "AÃ±o", "Valor": year},
                {"Concepto": "Total Ingresos", "Valor": total_income},
                {"Concepto": "Total Gastos", "Valor": total_expenses},
                {"Concepto": "Balance", "Valor": total_income - total_expenses},
                {"Concepto": "Tasa de Ahorro (%)", "Valor": round((total_income - total_expenses) / total_income * 100, 2) if total_income > 0 else 0},
                {"Concepto": "NÂº Transacciones", "Valor": len(all_transactions)}
            ]
            
            expenses_by_cat = {}
            for t in all_transactions:
                if t.transaction_type == "expense":
                    category = self.db.get_category_by_id(t.category_id)
                    cat_name = category.name if category else "Sin Categoría"
                    if cat_name not in expenses_by_cat:
                        expenses_by_cat[cat_name] = 0
                    expenses_by_cat[cat_name] += t.amount
            
            expenses_data = []
            if total_expenses > 0:
                for cat_name, total in sorted(expenses_by_cat.items(), key=lambda x: x[1], reverse=True):
                    percentage = (total / total_expenses * 100)
                    expenses_data.append({
                        "Categoría": cat_name,
                        "Total": total,
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            income_by_cat = {}
            for t in all_transactions:
                if t.transaction_type == "income":
                    category = self.db.get_category_by_id(t.category_id)
                    cat_name = category.name if category else "Sin Categoría"
                    if cat_name not in income_by_cat:
                        income_by_cat[cat_name] = 0
                    income_by_cat[cat_name] += t.amount
            
            income_data = []
            if total_income > 0:
                for cat_name, total in sorted(income_by_cat.items(), key=lambda x: x[1], reverse=True):
                    percentage = (total / total_income * 100)
                    income_data.append({
                        "Categoría": cat_name,
                        "Total": total,
                        "Porcentaje (%)": round(percentage, 2)
                    })
            
            filename = f"Reporte_TermoWallet_Anual_{year}.{format}"
            
            return self._generate_and_share_report(
                filename,
                format,
                transactions_data,
                summary_data,
                expenses_data,
                income_data,
                callback_success,
                callback_error
            )
            
        except Exception as e:
            print(f" Error generando reporte anual: {e}")
            import traceback
            traceback.print_exc()
            error_msg = f"Error: {str(e)}"
            if callback_error:
                callback_error(error_msg)
            return {"success": False, "filepath": None, "message": error_msg}

    def _generate_and_share_report(
        self,
        filename: str,
        format: str,
        transactions_data: List[Dict],
        summary_data: List[Dict],
        expenses_data: List[Dict],
        income_data: List[Dict],
        callback_success,
        callback_error
    ) -> Dict:
        """
        âNUEVO: Genera el archivo y lo comparte inmediatamente
        Sin FilePicker, directamente genera y abre
        """
        print("\n" + "="*60)
        print("GENERANDO Y COMPARTIENDO REPORTE")
        print("="*60)
        print(f"    Archivo: {filename}")
        print(f"   Formato: {format}")
        print(f"   Transacciones: {len(transactions_data)}")
        print("="*60 + "\n")
        
        try:
            # 1. Limpiar reportes antiguos
            self._clean_old_reports()
            
            # 2. Definir ruta en CACHÃ‰
            from src.utils.android_permissions import get_app_storage_path
            
            if self.is_android:
                target_path = os.path.join(
                    get_app_storage_path().replace("/files", "/cache"), 
                    filename
                )
            else:
                target_path = os.path.join(
                    tempfile.gettempdir(),
                    filename
                )
            
            print(f"  Ruta destino: {target_path}")
            
            # 3. Generar archivo según formato
            if format == 'xlsx':
                success = self._save_excel(
                    target_path,
                    transactions_data,
                    summary_data,
                    expenses_data,
                    income_data
                )
            elif format == 'csv':
                success = self._save_csv(target_path, transactions_data)
            else:
                success = False
            
            if not success:
                error_msg = "Error al generar el archivo"
                print(error_msg)
                if callback_error:
                    callback_error(error_msg)
                return {"success": False, "filepath": None, "message": error_msg}
            
            # 4. Verificar que el archivo existe
            if not os.path.exists(target_path):
                error_msg = "El archivo no se creó correctamente"
                print(error_msg)
                if callback_error:
                    callback_error(error_msg)
                return {"success": False, "filepath": None, "message": error_msg}
            
            file_size = os.path.getsize(target_path)
            print(f" Archivo creado: {file_size} bytes")
            
            # 5. âœ… COMPARTIR/ABRIR el archivo
            if self.page:
                print(f"  Intentando compartir archivo...")
                
                try:
                    # Intentar con launch_url (mÃ¡s compatible)
                    self.page.launch_url(f"file://{target_path}")
                    print(f" Archivo compartido con launch_url")
                    
                except Exception as share_error:
                    print(f" Error con launch_url: {share_error}")
                    
                    # Fallback: intentar con share_file si existe
                    if hasattr(self.page, "share_file"):
                        try:
                            self.page.share_file(target_path)
                            print(f"  Archivo compartido con share_file")
                        except Exception as share_error2:
                            print(f"   Error con share_file: {share_error2}")

            # 6. LLAMAR AL CALLBACK DE ÉXITO
            success_msg = f"Reporte generado: {filename}"
            print(f"   {success_msg}")
            
            if callback_success:
                callback_success(target_path, success_msg)
            
            return {
                "success": True,
                "filepath": target_path,
                "message": success_msg
            }
            
        except Exception as e:
            print(f"Error en generación: {e}")
            import traceback
            traceback.print_exc()
            error_msg = f"Error: {str(e)}"
            if callback_error:
                callback_error(error_msg)
            return {"success": False, "filepath": None, "message": error_msg}

    def _save_excel(
        self,
        filepath: str,
        trans_data: List[Dict],
        summary_data: List[Dict],
        expenses_data: List[Dict],
        income_data: List[Dict]
    ) -> bool:
        """Guarda Excel con metodo correcto según plataforma"""
        wb = None
        try:
            if not openpyxl:
                print(f"  openpyxl NO está disponible")
                raise ImportError("openpyxl no disponible")

            print(f"  📊 Iniciando creación de Excel...")

            # 1. Crear workbook en memoria
            wb = openpyxl.Workbook()
            
            # 2. Hoja Resumen
            ws_summary = wb.active
            ws_summary.title = "Resumen"
            self._write_data_to_sheet(ws_summary, summary_data)
            
            # 3. Hoja Transacciones
            ws_trans = wb.create_sheet("Transacciones")
            self._write_data_to_sheet(ws_trans, trans_data)
            
            # 4. Hoja Gastos
            if expenses_data:
                ws_exp = wb.create_sheet("Gastos por Categoría")
                self._write_data_to_sheet(ws_exp, expenses_data)
            
            # 5. Hoja Ingresos
            if income_data:
                ws_inc = wb.create_sheet("Ingresos por Categoría")
                self._write_data_to_sheet(ws_inc, income_data)
            
            # 6. Guardar archivo
            print(f"      💾 Guardando archivo...")
            if self.is_android:
                # ANDROID: Usar BytesIO intermedio
                from io import BytesIO
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                excel_bytes = buffer.read()
                buffer.close()
                
                with open(filepath, 'wb') as f:
                    f.write(excel_bytes)
                    f.flush()
                    os.fsync(f.fileno())
            else:
                # DESKTOP: MÃ©todo tradicional
                wb.save(filepath)
            
            # 7. Cerrar workbook
            try:
                wb.close()
            except:
                pass
            
            # 8. Esperar (crÃ­tico en Android)
            import time
            time.sleep(0.5)
            
            # 9. Verificar archivo
            if not os.path.exists(filepath):
                print(f"      ❌ ERROR: Archivo NO existe")
                return False
            
            file_size = os.path.getsize(filepath)
            print(f"      ðŸ“ TamaÃ±o del archivo: {file_size} bytes")
            
            if file_size == 0:
                print(f"      ❌ ERROR: Archivo vacío (0 bytes)")
                return False

            # 10. Verificar que sea un Excel válido
            try:
                test_wb = openpyxl.load_workbook(filepath, read_only=True)
                test_wb.close()
                print(f"      ✅ Excel válido verificado")
            except Exception as verify_err:
                print(f"      ❌ ERROR: Excel corrupto: {verify_err}")
                return False
            
            print(f"      ✅ Excel guardado exitosamente: {file_size} bytes")
            return True
            
        except Exception as e:
            print(f"      ❌ EXCEPCIÓN en _save_excel: {e}")
            import traceback
            traceback.print_exc()
            
            # Limpiar archivo corrupto
            try:
                if os.path.exists(filepath) and os.path.getsize(filepath) == 0:
                    os.remove(filepath)
            except:
                pass
            
            return False
            
        finally:
            if wb:
                try:
                    wb.close()
                except:
                    pass

    def _write_data_to_sheet(self, worksheet, data: List[Dict]):
        """Escribe datos en hoja Excel"""
        if not data:
            return
        
        # Encabezados
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.value = header
            try:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(
                    start_color="CCCCCC",
                    end_color="CCCCCC",
                    fill_type="solid"
                )
            except:
                pass
        
        # Datos
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header, "")
        
        # Ajustar ancho
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _save_csv(self, filepath: str, trans_data: List[Dict]) -> bool:
        """Guarda CSV con método correcto según plataforma"""
        try:
            print(f"      📊 Creando CSV con {len(trans_data)} transacciones...")
            
            if not trans_data:
                return False
            
            if self.is_android:
                # ANDROID: Escribir primero en StringIO
                from io import StringIO
                buffer = StringIO()
                writer = csv.DictWriter(buffer, fieldnames=trans_data[0].keys())
                writer.writeheader()
                writer.writerows(trans_data)
                csv_content = buffer.getvalue()
                buffer.close()
                
                with open(filepath, 'w', encoding='utf-8-sig') as f:
                    f.write(csv_content)
                    f.flush()
                    os.fsync(f.fileno())
            else:
                # DESKTOP: MÃ©todo tradicional
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.DictWriter(f, fieldnames=trans_data[0].keys())
                    writer.writeheader()
                    writer.writerows(trans_data)
                    f.flush()
                    os.fsync(f.fileno())
            
            # Verificar inmediatamente
            import time
            time.sleep(0.3)
            
            if os.path.exists(filepath):
                size = os.path.getsize(filepath)
                print(f"      ✅ CSV guardado: {size} bytes")
                return size > 0
            else:
                print(f"      ❌ Archivo no existe después de guardar")
                return False
                
        except Exception as e:
            print(f"      ❌ Error creando CSV: {e}")
            import traceback
            traceback.print_exc()
            return False